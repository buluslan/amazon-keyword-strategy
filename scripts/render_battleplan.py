"""作战图渲染器。

输入：Agent 分析好的作战图数据（JSON），含 verdict / keywords[] / negatives[] / data_sources
输出：output/{ASIN}_{日期}/ 下 作战图.xlsx + 作战图.md + 词表.json

Agent 负责语义判断（标品判别/词根分类/匹配建议/打法），产出结构化 JSON；
数值分档（ABA 层级/竞争打分/PPC 四象限）由本脚本按固定界限确定性计算——
同一输入永远同一输出，可复现。本脚本负责渲染 + 数值分档。

渲染规则钉死（battleplan-template.md §5）：
- 字段缺失统一标 [缺失]，不用空格/破折号/未知
- 未实测/估算值带 ⏳ 标记
- 不脑补数字（没有就 [缺失]）
- 去重（同词一行）
- 排序（精准词表按搜索量降序，缺失排末尾）
"""

import argparse
import datetime
import json
import logging
import os
import re
from typing import Dict, List, Optional

from word_frequency import tokenize

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

_OUTPUT_ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")

# 精准词表列定义：(JSON键, 列标题)
PRECISION_COLS = [
    ("keyword", "关键词"),
    ("search_volume", "搜索量"),
    ("aba_rank", "ABA排名"),
    ("aba_tier", "ABA层级"),
    ("competition_score", "竞争分"),
    ("competition_level", "竞争强度"),
    ("word_root", "词根类型"),
    ("match_type", "建议匹配"),
    ("action", "建议动作"),
    ("stage", "阶段性"),
]
NEGATIVE_COLS = [
    ("keyword", "关键词"),
    ("root", "所属否定词根"),
    ("reason", "否定原因"),
    ("negate_type", "否定类型"),
]
ROOT_SUMMARY_COLS = [
    ("token", "词根"),
    ("attribute_dim", "属性维度"),
    ("keywords_count", "命中关键词数"),
    ("search_volume_sum", "搜索量合计"),
    ("volume_share", "搜索量占比"),
    ("comp_rank_pct", "竞争分位"),
    ("opp_score", "机会分"),
    ("sample_keywords", "示例词"),
]
CAMPAIGN_GROUP_COLS = [
    ("group", "广告组"),
    ("keywords_count", "词数"),
    ("match_hint", "匹配倾向"),
    ("bid_hint", "竞价倾向"),
    ("skeleton_desc", "组级说明"),
    ("budget_hint", "预算倾向(注入)"),
    ("launch_advice", "开组建议(注入)"),
    ("sample_keywords", "代表词(Top10)"),
]

# 字段说明 sheet：机制说明写死（跨任务稳定），运行时口径由 payload["data_notes"] 注入
FIELD_GUIDE = [
    ("keyword", "关键词", "原始关键词，去重后保留"),
    ("search_volume", "搜索量", "0 = 数据源返回的真实零搜索（词真实存在但本周期无搜索量），与 [缺失]（源未提供）含义不同，不混用"),
    ("aba_rank", "ABA排名", "ABA 搜索频率排名原始值（全站统一绝对标尺，跨品类可比），源未返回标 [缺失]"),
    ("aba_tier", "ABA层级", "固定界限切档（脚本计算，可复现）：ABA排名 ≤1万 大词 / 1万~10万 中词 / 10万~50万 小词 / >50万 长尾（界限为行业经验锚点）。无 ABA排名时按本表搜索量分位兜底：≥P80 大 / ≥P50 中 / ≥P20 小 / 其余长尾"),
    ("competition_score", "竞争分", "0-100 确定性打分：本次可用竞争指标等权均值×100。0-1 型指标（点击/转化集中度、占位率）直接用原值，占位率族先组内平均；非归一指标（SPR/商品数）按本表 min-max 归一。本次实际依据的指标清单见下方口径说明"),
    ("competition_level", "竞争强度", "固定切点（脚本计算，可复现）：竞争分 ≥70 高 / 40~69 中 / <40 低。源独家指标（如占位率）作为竞争分输入参与判定，不单独成列，依据由口径说明交代"),
    ("word_root", "词根类型", "六分类：共性（最基础叫法）/ 属性（修饰维度）/ 品牌 / 品类（其他叫法）/ 受众（人群场景）/ 否定。属性词带维度后缀「属性·<维度>」，维度从本次词库现场归纳（如 有线/接口/尺寸），非预置清单"),
    ("品牌双分类", "品牌词双分类", "本品品牌词 → 主攻（归共性/属性处理，绝不能截流）；竞品品牌词 → 截流（精准，独立广告组）"),
    ("match_type", "建议匹配", "精准/词组/广泛，由竞争强度 + 词根类型推导"),
    ("action", "建议动作", "主攻/截流/广泛铺词/观察/否定，由词根类型 + 竞争强度推导"),
    ("stage", "阶段性", "词维度：该词适合在哪个阶段主打。强依据=趋势数据（不带⏳），弱依据=词规模推断（标⏳）；数据基础见下方口径说明"),
    ("ppc_quadrant", "PPC四象限", "ABA层级 × 竞争强度档的确定性交叉（脚本计算）：量高（大词/中词）竞低=蓝海主攻 / 量高竞高=精准控预算 / 量低竞低=长尾广泛 / 量低竞高=不建议打；任一档缺失归「数据不足」"),
    ("root", "所属否定词根", "同族否定词共享的词根（如 adapter 族）；同族优先词组否定——否一个词根即可批量拦截整族衍生长尾"),
    ("volume_share", "搜索量占比", "该词根族搜索量合计占本表（进入词根汇总的族）合计比例（脚本计算，本表内口径）"),
    ("comp_rank_pct", "竞争分位", "族内竞争分均值在进入词根汇总的族中的相对分位（0-100，越高=竞争越强；脚本计算，本表内相对，无绝对阈值）"),
    ("opp_score", "机会分", "0.5×流量分位 + 0.5×(100−竞争分位)，本表内相对位次加权（脚本计算，可复现；跨品类可比的是排序不是数值）。无搜索量字段时降级为词频口径，品牌词不参与聚合，见判读说明"),
    ("brand_role", "品牌标记", "词级可选注入（Agent 品牌双分类判断的落盘形态）：own=本品品牌 / competitor=竞品品牌 / eco=兼容生态。广告结构建议的品牌防御组按 own 聚合；本品品牌词 word_root 仍归共性/属性（主攻）"),
    ("campaign_groups", "广告结构建议", "六组确定性分组骨架（品牌防御/核心攻坚/精准长尾/竞品截流/场景拓展/否定清单，按品牌标记×动作×词根交叉）。竞价只给档位（高/中高/中/低），组级打法说明/预算倾向/开组顺序由 Agent 注入（两段式）——具体预算数字拆解与 Campaign 搭建执行不在本 skill 范围"),
]


def _field_guide_rows(payload: Dict) -> List[Dict]:
    """字段说明数据：机制写死 + 运行时口径（data_notes 可选）合并。按 field 键定位注入。"""
    notes = payload.get("data_notes") or {}
    rows = [{"field": f, "meaning": m, "how": h} for f, m, h in FIELD_GUIDE]
    by_field = {r["field"]: r for r in rows}
    unit = notes.get("search_volume_unit")
    if unit and "search_volume" in by_field:
        by_field["search_volume"]["how"] += f"｜时间单位：{unit}（本次数据源口径）"
    basis = notes.get("competition_basis")
    if basis:
        for key in ("competition_score", "competition_level"):
            if key in by_field:
                by_field[key]["how"] += f"｜本次依据：{basis}"
    stage = notes.get("stage_basis")
    if stage and "stage" in by_field:
        by_field["stage"]["how"] += f"｜{stage}"
    if notes.get("warning"):
        rows.append({"field": "⚠️ 预警", "meaning": "结构失配警示（render 自动计算）", "how": notes["warning"]})
    for n in notes.get("observations", []):
        rows.append({"field": "诊断要点", "meaning": "诊断要点/数据局限（无需确认的事实性说明）", "how": n})
    for n in notes.get("notes", []):
        rows.append({"field": "⏳ 待确认", "meaning": "推断/待核实项", "how": n})
    return rows


# 原始数据 sheet：判定字段（分析产物）排除，其余字段全量留底
# aba_tier/competition_score/competition_level 是判定产物（固定界限切档/打分），原始对应是 aba_rank/竞争原始指标
VERDICT_FIELDS = {"word_root", "match_type", "action", "stage", "ppc_quadrant",
                  "aba_tier", "competition_score", "competition_level", "brand_role"}


# ── 数值分档（确定性计算，界限写死可复现；语义判断归 Agent） ──────────────

# ABA 排名是全站统一绝对标尺，跨品类可比 → 固定界限（行业经验锚点）
ABA_TIER_BOUNDS = "≤1万 大词 / 1万~10万 中词 / 10万~50万 小词 / >50万 长尾"
COMPETITION_BOUNDS = "≥70 高 / 40~69 中 / <40 低"

# 竞争分指标槽：槽名 → 候选路径。0-1 型（ratio）直接用原值，多路径命中取均值；
# 非归一型（norm）按本表 min-max 归一后进入
_COMP_RATIO_SLOTS = {
    "Top3点击集中度": ["competition.click_concentration", "click_concentration", "top3_click_share"],
    "Top3转化集中度": ["competition.conv_concentration", "conv_concentration"],
    "占位率(族内均值)": [
        "relevance.coverage.top4", "relevance.coverage.top8", "relevance.coverage.top16",
        "relevance.coverage.top32", "relevance.coverage.top48",
        "coverage.top4", "coverage.top8", "coverage.top16", "coverage.top32", "coverage.top48",
        "coverage_ratio",
    ],
}
_COMP_NORM_SLOTS = {
    "SPR": ["competition.spr", "spr"],
    "商品数": ["competition.products", "products"],
}


def _dig(obj, path: str):
    """按点分路径取嵌套值，取不到返回 None。"""
    cur = obj
    for part in path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
    return cur


def _pctile(sorted_vals: List, p: int):
    """简单分位：有序列表取第 p 百分位位置值（不插值）。"""
    return sorted_vals[min(len(sorted_vals) - 1, len(sorted_vals) * p // 100)]


def _aba_tier(kw: Dict, vol_pcts) -> Optional[str]:
    """ABA 层级：aba_rank 优先固定界限切档；无排名按本表搜索量分位兜底（同为确定性规则）。"""
    rank = kw.get("aba_rank")
    if isinstance(rank, (int, float)):
        if rank <= 10000:
            return "大词"
        if rank <= 100000:
            return "中词"
        if rank <= 500000:
            return "小词"
        return "长尾"
    v = kw.get("search_volume")
    if isinstance(v, (int, float)) and vol_pcts:
        p80, p50, p20 = vol_pcts
        if v >= p80:
            return "大词"
        if v >= p50:
            return "中词"
        if v >= p20:
            return "小词"
        return "长尾"
    return None


def _competition_scores(keywords: List[Dict]) -> List[str]:
    """竞争分（0-100）：可用指标槽等权均值×100，就地写 competition_score / competition_level。

    返回本次实际命中的指标槽名（供 data_notes.competition_basis 自动交代口径）。
    指标全缺的词：两字段写 None（json 落 null，渲染为 [缺失]）。
    """
    extracted = []  # [(kw, {槽名: 原始值})]
    slot_hit: Dict[str, int] = {}
    for kw in keywords:
        vals = {}
        for slot, paths in _COMP_RATIO_SLOTS.items():
            hits = [v for v in (_dig(kw, p) for p in paths)
                    if isinstance(v, (int, float)) and 0 <= v <= 1]
            if hits:
                vals[slot] = sum(hits) / len(hits)
                slot_hit[slot] = slot_hit.get(slot, 0) + 1
        for slot, paths in _COMP_NORM_SLOTS.items():
            for p in paths:
                v = _dig(kw, p)
                if isinstance(v, (int, float)):
                    vals[slot] = float(v)
                    slot_hit[slot] = slot_hit.get(slot, 0) + 1
                    break
        extracted.append((kw, vals))
    # 非归一槽本表 min-max
    norm_range = {}
    for slot in _COMP_NORM_SLOTS:
        vs = [vals[slot] for _, vals in extracted if slot in vals]
        if vs:
            norm_range[slot] = (min(vs), max(vs))
    for kw, vals in extracted:
        if not vals:
            kw["competition_score"] = None
            kw["competition_level"] = None
            continue
        parts = []
        for slot, v in vals.items():
            if slot in norm_range:
                lo, hi = norm_range[slot]
                v = 0.5 if hi <= lo else (v - lo) / (hi - lo)
            parts.append(v)
        score = round(100 * sum(parts) / len(parts))
        kw["competition_score"] = score
        kw["competition_level"] = "高" if score >= 70 else ("中" if score >= 40 else "低")
    return [f"{s}({n}词)" for s, n in slot_hit.items()]


def _apply_deterministic_verdicts(payload: Dict) -> None:
    """数值分档三件套（ABA 层级 / 竞争打分 / PPC 四象限）——覆盖写入，确定性优先。

    Agent 若手填了这三个档位字段也会被此处的固定规则计算结果覆盖：
    同一输入永远同一输出，消灭逐次漂移。语义判断（词根分类/判别/打法）不在此层。
    """
    keywords = payload.get("keywords") or []
    # 分位池排除真实零搜索（0 词真实存在但本周期无搜索，进池会压低分位阈值）
    vols = sorted(kw["search_volume"] for kw in keywords
                  if isinstance(kw.get("search_volume"), (int, float))
                  and kw["search_volume"] > 0)
    vol_pcts = tuple(_pctile(vols, p) for p in (80, 50, 20)) if vols else None
    for kw in keywords:
        kw["aba_tier"] = _aba_tier(kw, vol_pcts)
    comp_used = _competition_scores(keywords)
    for kw in keywords:
        tier, comp = kw.get("aba_tier"), kw.get("competition_level")
        if tier and comp:
            vol_high = tier in ("大词", "中词")
            kw["ppc_quadrant"] = ("蓝海主攻" if vol_high and comp != "高" else
                                  "精准控预算" if vol_high else
                                  "长尾广泛" if comp != "高" else "不建议打")
        else:
            kw["ppc_quadrant"] = None
    # 口径自动交代：脚本实际用了哪些指标槽（覆盖 Agent 手写，保证口径与计算一致）
    if comp_used:
        notes = _ensure_notes(payload)
        notes["competition_basis"] = (
            f"竞争分 = {' + '.join(comp_used)}（等权均值×100）；档位固定切点：{COMPETITION_BOUNDS}"
        )
RAW_FIELD_CN = {
    "keyword": "关键词", "search_volume": "搜索量", "aba_rank": "ABA排名",
    "click_concentration": "点击集中度", "products": "商品数", "spr": "SPR",
    "conversion_rate": "转化率", "purchase_rate": "购买率", "bid": "建议竞价",
    "relevance": "相关性", "coverage_ratio": "覆盖率/占位率", "trend": "趋势",
    "traffic_percentage": "流量占比", "rank_position": "自然排名", "ad_position": "广告排名",
    # 平铺子列（object 字段展开）
    "relevance.tier": "相关性档", "relevance.score": "相关性得分",
    "coverage.top4": "Top4占位率", "coverage.top8": "Top8占位率",
    "coverage.top16": "Top16占位率", "coverage.top32": "Top32占位率", "coverage.top48": "Top48占位率",
    "competition.click_concentration": "Top3点击集中度", "competition.conv_concentration": "Top3转化集中度",
    "competition.spr": "SPR", "competition.products": "商品数",
    "competition.ad_products": "广告竞品数", "competition.supply_demand_ratio": "供需比",
    "_translation": "中文翻译", "_volume_period": "搜索量周期", "_source": "数据源",
    "source_asin": "反查来源ASIN",
}


def _raw_sheet(payload: Dict):
    """原始数据留底：动态列 = 全部关键词的字段并集 - 判定字段。

    object 字段（relevance/competition 等）平铺成子列（relevance.tier / coverage.top4），
    不整坨 str() 渲染。词表存在任一增强字段才返回 (headers, rows)；纯关键词输入返回 None。
    """
    keywords = payload.get("keywords") or []
    # 存在 dict 值的顶层 key（平铺目标）：本行该 key 为 None 时不输出顶层幽灵列，
    # 否则同一逻辑字段出现两条表示路径（competition 平铺子列 + 全 [缺失] 的裸 competition 列）
    dict_top_keys = {k for kw in keywords if isinstance(kw, dict)
                     for k, v in kw.items() if isinstance(v, dict)}

    def flat(kw):
        """object 平铺一层嵌套（coverage 再下一层），返回 {平铺key: 值}。

        list 值（如 source_asin 反查来源数组）转逗号连接字符串展示。
        """
        out = {}
        for k, v in kw.items():
            if k in VERDICT_FIELDS:
                continue
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    if isinstance(v2, dict):
                        for k3, v3 in v2.items():
                            out[f"{k2}.{k3}"] = v3
                    else:
                        out[f"{k}.{k2}"] = v2
            elif v is None and k in dict_top_keys:
                continue
            elif isinstance(v, list):
                out[k] = ", ".join(str(x) for x in v) if v else None
            else:
                out[k] = v
        return out

    flat_rows = [flat(kw) for kw in keywords]
    all_keys: List[str] = []
    for r in flat_rows:
        for k in r.keys():
            if k not in all_keys:
                all_keys.append(k)
    # source_asin 是结构性补齐字段（_fill_null_keys 强制补 null），不算增强字段——
    # 纯关键词输入不应因它出现"原始数据"幽灵 sheet
    if not any(k != "keyword" and k != "source_asin" and not k.startswith("_") for k in all_keys):
        return None  # 纯关键词输入：无原始数据可留
    headers = [RAW_FIELD_CN.get(k, k) for k in all_keys]
    rows = [[_cell(r.get(k)) for k in all_keys] for r in flat_rows]
    return headers, rows


def _cell(val) -> str:
    """渲染单个单元格值：None/空 → [缺失]；list（如 sample_keywords）→ 逗号连接。"""
    if val is None or val == "" or val == []:
        return "[缺失]"
    if isinstance(val, list):
        return ", ".join(str(x) for x in val)
    return str(val)


def _ensure_notes(payload: Dict) -> Dict:
    """取 data_notes（显式 null 也当空 dict 重置后回写——setdefault 不防 null）。"""
    notes = payload.get("data_notes") or {}
    payload["data_notes"] = notes
    return notes


def _sort_keywords(keywords: List[Dict]) -> List[Dict]:
    """按搜索量降序，缺失排末尾。"""
    def keyf(kw):
        v = kw.get("search_volume")
        return (0, -v) if isinstance(v, (int, float)) else (1, 0)
    return sorted(keywords, key=keyf)


def _ppc_quadrant(keywords: List[Dict]) -> Dict[str, List[Dict]]:
    """按 ppc_quadrant 字段分组（字段由 _apply_deterministic_verdicts 确定性算好：
    ABA层级 × 竞争强度档交叉）。未算出（任一档缺失）归"数据不足"。
    """
    buckets = {"蓝海主攻": [], "精准控预算": [], "长尾广泛": [], "不建议打": [], "数据不足": []}
    for kw in keywords:
        q = kw.get("ppc_quadrant")
        buckets[q if q in buckets else "数据不足"].append(kw)
    return buckets


def _rank_pct_map(keys_vals: Dict[str, Optional[float]]) -> Dict[str, Optional[float]]:
    """值 → 本表内相对分位（0-100）。同值同分位（取同值组最小名次）。

    纯相对位次、零绝对阈值——跨品类可比的是排序，不是数值。
    少于 2 个有效值时全部 None（单值无分位意义）。
    """
    valid = {k: v for k, v in keys_vals.items() if isinstance(v, (int, float))}
    out: Dict[str, Optional[float]] = {k: None for k in keys_vals}
    if len(valid) < 2:
        return out
    items = sorted(valid.items(), key=lambda x: x[1])
    n = len(items)
    starts: Dict[float, int] = {}
    for i, (_, v) in enumerate(items):
        starts.setdefault(v, i)
    for i, (k, v) in enumerate(items):
        out[k] = round(100 * starts[v] / (n - 1))
    return out


def _root_summary(payload: Dict, top_n: int = 30) -> List[Dict]:
    """词根汇总（Sheet 4 数据）：token 级聚合 + 机会分，全部确定性计算。

    维度列取 payload["root_dims"]（Agent 归纳的 token→维度映射，可选注入），
    未提供时标 [缺失]——维度归纳是判断类动作归 Agent，render 只做机械聚合
    （措辞边界：判断归 Agent，聚合归脚本）。

    机会分三列（volume_share / comp_rank_pct / opp_score）全部本表内相对口径：
    无搜索量字段时流量维度降级为词频分位（口径写入 data_notes.opp_score_basis）。
    """
    keywords = [kw for kw in payload.get("keywords") or [] if kw.get("keyword")]
    root_dims = payload.get("root_dims") or {}
    token_kws: Dict[str, List[Dict]] = {}
    for kw in keywords:
        # 品牌词不参与词根汇总/属性机会聚合：token 级聚合会把品牌 compound 的流量
        # 卷进泛称 token 行（如「litter robot 4」污染 robot 行），顶掉真实主攻簇判断——
        # 品牌流量看品牌防御组/竞品截流组，不掺需求结构视图。
        # 判定双条件：word_root=品牌（竞品/兼容生态）或 brand_role 已标记（本品品牌词
        # word_root 归共性/属性，靠 brand_role=own 识别）
        if str(kw.get("word_root") or "") == "品牌" or kw.get("brand_role"):
            continue
        for t in set(tokenize(kw["keyword"])):
            token_kws.setdefault(t, []).append(kw)
    rows = []
    # 同频 token 按字典序 tie-break：set(tokenize()) 的迭代序受进程 hash 随机化影响，
    # 不加次键会出现同输入不同输出（可复现性缺陷）
    top_tokens = {t for t, _ in sorted(token_kws.items(), key=lambda x: (-len(x[1]), x[0]))[:top_n]}
    # 保底：已归纳维度的 token 不受 Top 截断（小频次维度若被截掉，该维度在
    # 词根汇总/属性机会视图整体消失——维度可见性优先于行数控制）
    dim_tokens = {t for t in root_dims if t in token_kws}
    for token in sorted(top_tokens | dim_tokens):
        kws = token_kws[token]
        vols = [kw.get("search_volume") for kw in kws
                if isinstance(kw.get("search_volume"), (int, float))]
        comps = [kw.get("competition_score") for kw in kws
                 if isinstance(kw.get("competition_score"), (int, float))]
        rows.append({
            "token": token,
            "attribute_dim": root_dims.get(token, "[缺失]"),
            "keywords_count": len(kws),
            "search_volume_sum": sum(vols) if vols else "[缺失]",
            "_vol_sum": sum(vols) if vols else None,     # 机会分内部用（json 落盘前剔除）
            "_comp_avg": (sum(comps) / len(comps)) if comps else None,
            "sample_keywords": [kw["keyword"] for kw in kws[:3]],  # json 落数组（§6），xlsx 渲染时 join
        })
    # ── 机会分（确定性，本表相对位次） ──────────────────────────────
    vol_total = sum(r["_vol_sum"] for r in rows if isinstance(r["_vol_sum"], (int, float)))
    has_vol = vol_total > 0
    vol_rank = _rank_pct_map({r["token"]: (r["_vol_sum"] if has_vol else r["keywords_count"]) for r in rows})
    comp_rank = _rank_pct_map({r["token"]: r["_comp_avg"] for r in rows})
    for r in rows:
        vr, cr = vol_rank.get(r["token"]), comp_rank.get(r["token"])
        r["volume_share"] = (round(100 * r["_vol_sum"] / vol_total, 1)
                             if has_vol and isinstance(r["_vol_sum"], (int, float)) else "[缺失]")
        r["comp_rank_pct"] = cr if cr is not None else "[缺失]"
        if vr is None:
            r["opp_score"] = "[缺失]"
        else:
            r["opp_score"] = round(0.5 * vr + 0.5 * (100 - cr)) if cr is not None else vr
    return rows


# ── 广告结构建议（确定性分组骨架 + Agent 注入槽，两段式） ────────────────
# 匹配/竞价是组级固定倾向（档位，不给数字预算）；组级打法说明/预算倾向/开组顺序
# 由 Agent 经 payload.campaign_groups 注入（判断归 Agent，聚合归脚本）
CAMPAIGN_GROUP_DEFS = [
    ("品牌防御组", "本品品牌词防御——自有品牌是核心流量，高竞价守住，绝不能省",
     "精准+词组", "高"),
    ("核心攻坚组", "共性主攻词分层推进——大词广泛+中词词组+小词精准，逐步吃下核心流量",
     "大词广泛+中词词组+小词精准", "中"),
    ("精准长尾组", "属性词精准长尾——「小词带大词」的载体，出单性价比最稳的一组",
     "精准为主", "中高"),
    ("竞品截流组", "竞品/兼容生态品牌词截流——独立广告组+单独预算，低竞价捡量，严控 ACOS",
     "精准（独立组）", "低"),
    ("场景拓展组", "品类/受众词铺场景——广泛+词组探需求，跑出转化数据再升级打法",
     "广泛+词组", "中低"),
    ("否定清单", "挂到自动广告与广泛组的否定清单——词根族优先词组否定，拦整族长尾",
     "挂自动/广泛组", "—"),
]


def _campaign_groups(payload: Dict) -> tuple:
    """广告结构建议分组骨架（确定性）：品牌标记×动作×词根交叉分组。

    brand_role=own → 品牌防御组（Agent 品牌双分类判断的落盘形态，词级可选注入；
    未注入时该组为空并保留占位说明）。Agent 预注入的 campaign_groups 说明字段
    （description/budget_hint/launch_advice）按组名合并，骨架字段覆盖写入。
    返回 (groups, ungrouped_count)。
    """
    keywords = [kw for kw in payload.get("keywords") or []
                if kw.get("keyword")]  # 非 dict 元素已在 render 入口统一剔除
    negatives = payload.get("negatives") or []
    buckets: Dict[str, List[Dict]] = {name: [] for name, *_ in CAMPAIGN_GROUP_DEFS}
    ungrouped = 0
    for kw in keywords:
        brand_role = str(kw.get("brand_role") or "").strip().lower()
        action = str(kw.get("action") or "")
        root = str(kw.get("word_root") or "")
        # 观察词（打法未定）与否定词不入主攻组；品牌 own 标记优先于动作
        if brand_role == "own" and action not in ("否定", "观察"):
            buckets["品牌防御组"].append(kw)
        elif action == "截流":
            buckets["竞品截流组"].append(kw)
        elif action == "主攻" and root == "共性":
            buckets["核心攻坚组"].append(kw)
        elif action == "主攻" and root.startswith("属性"):
            buckets["精准长尾组"].append(kw)
        elif action in ("主攻", "广泛铺词") and root in ("品类", "受众"):
            buckets["场景拓展组"].append(kw)
        else:
            ungrouped += 1
    injected = {g.get("group"): g for g in payload.get("campaign_groups") or []
                if isinstance(g, dict) and g.get("group")}
    out = []
    for name, skeleton, match_hint, bid_hint in CAMPAIGN_GROUP_DEFS:
        if name == "否定清单":
            count = len(negatives)
            sample = [n.get("keyword") for n in negatives[:10] if n.get("keyword")]
        else:
            kws = buckets[name]
            count = len(kws)
            sample = [kw["keyword"] for kw in _sort_keywords(kws)[:10]]
        inj = injected.get(name) or {}
        out.append({
            "group": name,
            "keywords_count": count,
            "match_hint": match_hint,
            "bid_hint": bid_hint,
            "skeleton_desc": skeleton,
            "description": inj.get("description") or None,
            "budget_hint": inj.get("budget_hint") or None,
            "launch_advice": inj.get("launch_advice") or None,
            "sample_keywords": sample,
        })
    return out, ungrouped


def write_json(payload: Dict, path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def write_xlsx(payload: Dict, path: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # Sheet 1 精准词表
    ws1 = wb.active
    ws1.title = "精准词表"
    ws1.append([c[1] for c in PRECISION_COLS])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1a237e")
        cell.font = Font(bold=True, color="FFFFFF")
    for kw in _sort_keywords(payload.get("keywords") or []):
        ws1.append([_cell(kw.get(k)) for k, _ in PRECISION_COLS])
    _autowidth(ws1)

    # Sheet 2 否定词表
    ws2 = wb.create_sheet("否定词表")
    ws2.append([c[1] for c in NEGATIVE_COLS])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a237e")
    for neg in payload.get("negatives") or []:
        ws2.append([_cell(neg.get(k)) for k, _ in NEGATIVE_COLS])
    _autowidth(ws2)

    # Sheet 3 PPC 策略（四象限）
    ws3 = wb.create_sheet("PPC策略")
    buckets = _ppc_quadrant(payload.get("keywords") or [])
    legend = {
        "蓝海主攻": "🟢 高搜索 + 低竞争，主攻（精准+加预算）",
        "精准控预算": "🟡 高搜索 + 高竞争，精准匹配控预算盯 ACOS",
        "长尾广泛": "🟡 低搜索 + 低竞争，可广泛/词组铺",
        "不建议打": "🔴 低搜索 + 高竞争（高转化例外）",
        "数据不足": "⏳ 搜索量或竞争缺失，待补数据",
    }
    for bucket, desc in legend.items():
        ws3.append([bucket, desc])
        for kw in buckets.get(bucket, []):
            ws3.append(["", _cell(kw.get("keyword"))])
        ws3.append([])
    _autowidth(ws3)

    # Sheet 4 词根汇总（族级聚合 + 机会分）
    ws4 = wb.create_sheet("词根汇总")
    ws4.append([c[1] for c in ROOT_SUMMARY_COLS])
    # 机会分三列写数字（Excel 内可排序/计算；缺失才落 [缺失] 文本），其余列维持 _cell 统一文本
    _num_cols = {"volume_share", "comp_rank_pct", "opp_score"}

    def _root_cell(row: Dict, k: str):
        v = row.get(k)
        if k in _num_cols and isinstance(v, (int, float)):
            return v
        return _cell(v)
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a237e")
    for row in payload.get("root_summary") or _root_summary(payload):
        ws4.append([_root_cell(row, k) for k, _ in ROOT_SUMMARY_COLS])
    _autowidth(ws4)

    # Sheet 5 广告结构建议（分组骨架 + 注入槽，两段式）
    ws5 = wb.create_sheet("广告结构建议")
    ws5.append([c[1] for c in CAMPAIGN_GROUP_COLS])
    for cell in ws5[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a237e")
    for row in payload.get("campaign_groups") or []:
        values = []
        for k, _ in CAMPAIGN_GROUP_COLS:
            v = row.get(k)
            if k == "skeleton_desc":  # 组级说明列：Agent 注入的 description 优先，无则落骨架句
                v = row.get("description") or row.get("skeleton_desc")
            values.append(_cell(v))
        ws5.append(values)
    _autowidth(ws5)

    # Sheet 6 字段说明（就地可查：机制写死 + 运行时口径注入）
    ws6 = wb.create_sheet("字段说明")
    ws6.append(["字段", "含义", "怎么判定的"])
    for cell in ws6[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a237e")
    for row in _field_guide_rows(payload):
        ws6.append([_cell(row.get(k)) for k in ("field", "meaning", "how")])
    _autowidth(ws6)

    # Sheet 7 原始数据留底（数据源反查路径才有；纯关键词输入不出现）
    raw = _raw_sheet(payload)
    raw_header_len = 0
    if raw:
        headers, rows = raw
        raw_header_len = len(headers)
        ws7 = wb.create_sheet("原始数据")
        ws7.append(headers)
        for cell in ws7[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a237e")
        for r in rows:
            ws7.append(r)
        _autowidth(ws7)
    wb.save(path)
    # 列数回读自检（写完读回数列数，防串列）：每 sheet max_column == 表头列数
    _verify_xlsx_columns(path, [
        ("精准词表", len(PRECISION_COLS)),
        ("否定词表", len(NEGATIVE_COLS)),
        ("PPC策略", 2),
        ("词根汇总", len(ROOT_SUMMARY_COLS)),
        ("广告结构建议", len(CAMPAIGN_GROUP_COLS)),
        ("字段说明", 3),
    ] + ([("原始数据", raw_header_len)] if raw_header_len else []))


def _verify_xlsx_columns(path: str, expected: List[tuple]) -> None:
    """xlsx 落盘后读回校验每 sheet 列数与表头一致（错位即抛错，不静默交付坏表）。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    problems = []
    for name, header_len in expected:
        ws = wb[name]
        if ws.max_column != header_len:
            problems.append(f"{name}: {ws.max_column} 列 ≠ 表头 {header_len} 列")
    wb.close()
    if problems:
        raise RuntimeError("xlsx 列数自检失败（疑似串列）：" + "；".join(problems))


def write_md(payload: Dict, path: str) -> None:
    lines: List[str] = []
    verdict = payload.get("verdict", "[缺失]")
    anchor = payload.get("name_anchor") or _name_anchor(payload)  # 标题与文件名同一兜底锚
    sources = payload.get("data_sources") or []
    lines.append(f"# 关键词作战图 · {anchor}")
    lines.append("")
    lines.append(f"**判别结论**：{verdict}")
    lines.append(f"**数据来源**：{', '.join(sources) if sources else '[未标注]'}")
    lines.append(f"**关键词数**：{len(payload.get('keywords') or [])}　**否定词数**：{len(payload.get('negatives') or [])}")
    lines.append("")
    notes = payload.get("data_notes") or {}
    if notes:
        lines.append("## 判读说明")
        lines.append("")
        lines.append(f"- **搜索量口径**：{notes.get('search_volume_unit', '[缺失]')}")
        lines.append(f"- **竞争强度依据**：{notes.get('competition_basis', '[缺失]')}")
        lines.append(f"- **阶段性基础**：{notes.get('stage_basis', '[缺失]')}")
        if notes.get("opp_score_basis"):
            lines.append(f"- **机会分口径**：{notes['opp_score_basis']}")
        if payload.get("campaign_groups") is not None:
            lines.append("- **广告分组口径**：品牌标记×动作×词根交叉分组，六组判定链见字段说明 sheet；"
                         "观察词与否定词不入主攻组")
        if notes.get("warning"):
            lines.append(f"- **⚠️ 预警**：{notes['warning']}")
        for n in notes.get("observations", []):
            lines.append(f"- **诊断要点**：{n}")
        for n in notes.get("notes", []):
            lines.append(f"- **⏳ 待确认**：{n}")
        lines.append("")
    # 策略要点（Agent 产出的打法/广告组架构段，payload.strategy_summary 注入——
    # 同日重跑 render 覆盖 md 时该段随 payload 保留，不再依赖落盘后手工追加）
    strat = payload.get("strategy_summary")
    if strat:
        lines.append("## 策略要点")
        lines.append("")
        for s in ([strat] if isinstance(strat, str) else strat):
            lines.append(str(s))
            lines.append("")
    # 属性机会 Top（词根汇总的机会分视图，确定性计算；解读走 strategy_summary）
    root_rows = [r for r in (payload.get("root_summary") or [])
                 if isinstance(r.get("opp_score"), (int, float))]
    dim_rows = [r for r in root_rows
                if str(r.get("attribute_dim") or "") not in ("", "[缺失]", "None")]
    if dim_rows:
        top = sorted(dim_rows, key=lambda r: (-r["opp_score"], r["token"]))[:15]
        # 保底：每个已归纳维度至少一行代表，防小频次维度被 Top 截断后不可见
        seen = {r.get("attribute_dim") for r in top}
        for r in sorted(dim_rows, key=lambda r: (-r["opp_score"], r["token"])):
            if r.get("attribute_dim") not in seen:
                top.append(r)
                seen.add(r.get("attribute_dim"))
        lines.append("## 属性机会 Top")
        lines.append("")
        lines.append("> 机会分 = 流量分位与竞争优势的本表内相对加权（口径见判读说明）——排序可比，数值不跨表比；"
                     "仅列已归纳维度的属性词根，品牌词不参与。")
        lines.append("")
        lines.append("| 词根 | 属性维度 | 机会分 | 依据 |")
        lines.append("|---|---|---|---|")
        for r in top:
            vs, cr = r.get("volume_share"), r.get("comp_rank_pct")
            parts = [f"搜索量占比 {vs}%" if isinstance(vs, (int, float)) else "",
                     f"竞争分位 {cr}" if isinstance(cr, (int, float)) else ""]
            basis = " · ".join(p for p in parts if p) or "[依据缺失]"
            lines.append(f"| {r.get('token')} | {r.get('attribute_dim', '[缺失]')} | "
                         f"{r['opp_score']} | {basis} |")
        lines.append("")
    elif root_rows:
        lines.append("## 属性机会 Top")
        lines.append("")
        lines.append("> 未归纳出属性维度（root_dims 未注入）——属性机会视图待 Step 3 维度归纳后可见。")
        lines.append("")
    # 广告结构建议（分组骨架 + Agent 注入槽）
    cg = payload.get("campaign_groups") or []
    if cg:
        lines.append("## 广告结构建议")
        lines.append("")
        lines.append("> 倾向级建议：竞价只给档位，具体预算数字拆解与 Campaign 搭建执行不在本 skill 范围。")
        lines.append("")
        lines.append("| 广告组 | 词数 | 匹配倾向 | 竞价倾向 | 组级说明 |")
        lines.append("|---|---|---|---|---|")
        for g in cg:
            desc = g.get("description") or g.get("skeleton_desc")
            lines.append(f"| {g.get('group')} | {g.get('keywords_count')} | {g.get('match_hint')} | "
                         f"{g.get('bid_hint')} | {desc} |")
        lines.append("")
        for g in cg:
            extra = []
            if g.get("budget_hint"):
                extra.append(f"**{g.get('group')}·预算倾向**：{g['budget_hint']}")
            if g.get("launch_advice"):
                extra.append(f"**{g.get('group')}·开组建议**：{g['launch_advice']}")
            for e in extra:
                lines.append(f"- {e}")
        if any(g.get("budget_hint") or g.get("launch_advice") for g in cg):
            lines.append("")
    if payload.get("campaign_ungrouped"):
        lines.append(f"_（另有 {payload['campaign_ungrouped']} 词未入组：观察词/数据不足词不进主攻组；"
                     f"品牌词若未标注 brand_role 也暂不入品牌防御组，待注入后归组——见精准词表）_")
        lines.append("")
    lines.append("> 字段缺失标 [缺失]；估算/未实测值带 ⏳。不脑补数字。")
    lines.append("")
    lines.append("## 精准词表")
    lines.append("")
    lines.append("| " + " | ".join(c[1] for c in PRECISION_COLS) + " |")
    lines.append("|" + "|".join("---" for _ in PRECISION_COLS) + "|")
    for kw in _sort_keywords(payload.get("keywords") or [])[:50]:  # md 前 50 行，全量看 xlsx
        lines.append("| " + " | ".join(_cell(kw.get(k)) for k, _ in PRECISION_COLS) + " |")
    total = len(payload.get("keywords") or [])
    if total > 50:
        date_lbl = payload.get("date") or datetime.date.today().strftime("%Y%m%d")
        lines.append(f"\n_（仅显示前 50 行，全 {total} 词见 关键词作战图_{anchor}_{date_lbl}.xlsx）_")
    lines.append("")
    if payload.get("negatives"):
        lines.append("## 否定词表")
        lines.append("")
        lines.append("| " + " | ".join(c[1] for c in NEGATIVE_COLS) + " |")
        lines.append("|" + "|".join("---" for _ in NEGATIVE_COLS) + "|")
        for neg in (payload.get("negatives") or []):
            lines.append("| " + " | ".join(_cell(neg.get(k)) for k, _ in NEGATIVE_COLS) + " |")
        lines.append("")
    if verdict == "非标品":
        lines.append("## 非标品提示")
        lines.append("")
        lines.append("判别为非标品：出单词分散，精准单点打法效率低。本作战图为基础打法（词根分类 + 匹配建议 + 否定词），匹配建议整体偏广泛/词组铺长尾。")
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _autowidth(ws) -> None:
    """简易列宽自适应。"""
    for col in ws.columns:
        length = max(len(str(c.value)) for c in col if c.value is not None)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 4, 10), 40)


def _fill_null_keys(keywords: List[Dict]) -> None:
    """增强字段缺失的词补 key=null（§6：不省略 key，下游脚本/schema 校验可判空）。

    字段集 = 全部词出现过的 key 并集（源无关，Agent 产出哪个字段就补齐哪个）
    + 结构性字段 source_asin（反查来源；词表输入全库皆无并集不会包含，
    但作为场景标识字段强制补齐，与顶层 competitor_asins 空数组口径一致）。
    """
    all_keys: List[str] = ["source_asin"]
    for kw in keywords:
        for k in kw.keys():
            if k not in all_keys:
                all_keys.append(k)
    for kw in keywords:
        for k in all_keys:
            kw.setdefault(k, None)


def _norm_competitor_asins(val) -> List[str]:
    """competitor_asins 容错归一：字符串/None/列表 → 元素 upper、去空、去重保序的数组。"""
    if val is None:
        return []
    vals = [val] if isinstance(val, str) else (val if isinstance(val, list) else [])
    seen: Dict[str, None] = {}
    for a in vals:
        if a is None or a == "":
            continue
        s = str(a).strip().upper()
        if s:
            seen.setdefault(s, None)
    return list(seen.keys())


def _name_anchor(payload: Dict) -> str:
    """命名锚点逐级兜底：本品 ASIN > 词表名 > 核心关键词 > 首个竞品 ASIN > "词表"。

    竞品反查建库场景本品可能未上架/未提供 ASIN，文件名仍需自描述。
    每级 or 短路处理显式 null；锚点做文件名清洗（非法字符→-，空格→连字符），
    目录/文件/md 标题三处共用同一锚点。
    """
    competitor_asins = _norm_competitor_asins(payload.get("competitor_asins"))
    raw = (payload.get("asin") or payload.get("name") or payload.get("core_keyword")
           or (competitor_asins[0] if competitor_asins else None) or "词表")
    cleaned = re.sub(r'[/\\:*?"<>|]', "-", str(raw)).strip().replace(" ", "-")
    return cleaned or "词表"


def render(payload: Dict, output_dir: Optional[str] = None) -> str:
    """渲染作战图三件套，返回输出目录路径。"""
    # 脏输入防崩 + json 规范（无则空数组/空表，不省略 key）
    payload["keywords"] = payload.get("keywords") or []
    payload["negatives"] = payload.get("negatives") or []
    payload["competitor_asins"] = _norm_competitor_asins(payload.get("competitor_asins"))
    # 词级 source_asin 统一归一（upper/去重保序）——Agent 手写绕过 normalize/dedup 时兜底，
    # 大小写不归一会虚增竞品覆盖数
    for kw in payload["keywords"]:
        if isinstance(kw, dict) and kw.get("source_asin"):
            kw["source_asin"] = _norm_competitor_asins(kw["source_asin"])
    # 命名锚点（目录/文件名/md 标题共用）+ 回写 json（asin 字段保持真实值不造假）
    anchor = _name_anchor(payload)
    payload["name_anchor"] = anchor
    date = payload.get("date") or datetime.date.today().strftime("%Y%m%d")
    payload["date"] = date
    dirname = f"{anchor}_{date}"
    out = output_dir or os.path.join(_OUTPUT_ROOT, dirname)
    os.makedirs(out, exist_ok=True)

    # 脏输入统一清洗（入口一次，主路径与 Agent 手写 payload 同受保护）：
    # keywords/negatives 混入的非 dict 元素剔除；null 归一为空列表——
    # 主路径有过滤不算数，手写 payload 会绕过管线直达 render
    payload["keywords"] = [kw for kw in payload.get("keywords") or [] if isinstance(kw, dict)]
    payload["negatives"] = [n for n in payload.get("negatives") or [] if isinstance(n, dict)]

    # 数值分档先行（ABA 层级/竞争打分/四象限，覆盖写入 → json 落盘即含计算结果）
    _apply_deterministic_verdicts(payload)
    # 结构失配预警：否定占比 > 80% = 候选池与规格锚点/品牌归属严重失配，
    # 照常出图但必须警示（防用户拿 5 词作战图直接投放），口径由 render 自动写入
    # 预警用全量口径：negatives 是各族代表词（§6），全量在 negatives_raw——代表词取样会漏报
    n_kw = len(payload["keywords"])
    n_neg = len(payload.get("negatives_raw") or payload["negatives"] or [])
    if n_kw + n_neg > 0 and n_neg / (n_kw + n_neg) > 0.8:
        pct = round(100 * n_neg / (n_kw + n_neg))
        _ensure_notes(payload)["warning"] = (
            f"否定词占比 {n_neg}/{n_kw + n_neg}（{pct}%）——候选池与规格锚点/品牌归属严重失配："
            f"请核对锚点是否正确，或补充与本品同形态的对标竞品后再跑"
        )
    # 缺失 key 补 null（json 规范：不省略 key）+ 否定表按类型分组（词组优先，词根级否定优先）
    _fill_null_keys(payload["keywords"])
    payload["negatives"] = sorted(
        payload["negatives"],
        key=lambda n: {"词组": 0, "精准": 1}.get(str(n.get("negate_type")), 2),
    )
    # 词根汇总自动聚合写进 json（无条件覆盖，与数值分档"覆盖写入"哲学一致——预填作废）
    payload["root_summary"] = _root_summary(payload)
    # 机会分口径自动交代（render 覆盖写入，与 competition_basis 同模式）：
    # 全表无搜索量字段时流量维度降级词频口径——口径与计算一致，不靠 Agent 手填
    has_vol = any(isinstance(r.get("search_volume_sum"), (int, float))
                  for r in payload["root_summary"])
    notes = _ensure_notes(payload)
    notes["opp_score_basis"] = (
        "机会分 = 0.5×流量分位 + 0.5×(100−竞争分位)，本表内相对位次（无绝对阈值，跨表只比排序）；"
        if has_vol else
        "[机会分降级：无搜索量字段，流量维度按词频口径] "
        "机会分 = 流量分位，本表内相对位次（无绝对阈值，跨表只比排序）；"
    ) + "竞争分位缺失时仅按流量维度；品牌词不参与聚合（品牌流量看品牌防御/竞品截流组）"
    # root_summary 内部计算字段剔除（_vol_sum/_comp_avg 不进 json 交付物）
    for r in payload["root_summary"]:
        r.pop("_vol_sum", None)
        r.pop("_comp_avg", None)
    # 广告结构建议：确定性分组骨架（覆盖写入）+ Agent 预注入说明字段按组合并
    payload["campaign_groups"], ungrouped = _campaign_groups(payload)
    payload["campaign_ungrouped"] = ungrouped  # 未入组词数（观察/数据不足词），随 json 落盘
    label = f"{anchor}_{date}"  # 文件名自描述：类型打头 + 命名锚点 + 日期
    write_json(payload, os.path.join(out, f"关键词词表_{label}.json"))
    try:
        write_xlsx(payload, os.path.join(out, f"关键词作战图_{label}.xlsx"))
    except ImportError:
        logger.warning("openpyxl 未安装，跳过 xlsx，仅输出 md+json")
    write_md(payload, os.path.join(out, f"关键词作战图_{label}.md"))
    logger.info("作战图已落盘：%s（关键词作战图_%s.xlsx / .md + 关键词词表_%s.json）", out, label, label)
    return out


def main():
    parser = argparse.ArgumentParser(description="渲染关键词作战图（xlsx + md + json）")
    parser.add_argument("--in", dest="inp", required=True, help="Agent 分析好的作战图 JSON")
    parser.add_argument("--out", default=None, help="输出目录（默认 output/{ASIN}_{日期}/）")
    args = parser.parse_args()
    with open(args.inp, "r", encoding="utf-8") as f:
        payload = json.load(f)
    out = render(payload, args.out)
    print(f"✅ 作战图已生成：{out}")


if __name__ == "__main__":
    main()
